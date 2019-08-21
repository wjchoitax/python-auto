from selenium import webdriver
from openpyxl import Workbook, load_workbook
from os.path import exists
from datetime import datetime
import string

today = datetime.now()
monitor_path = '/Users/Alghost/Downloads/'
today_file_name = '%d_%d_%d.xlsx'%(today.year,today.month,today.day)
file_path = monitor_path + today_file_name

if exists(file_path):
    result_xlsx = load_workbook(file_path)
else:
    result_xlsx = Workbook()

worksheet = result_xlsx.active
worksheet['A2'] = '최근 가격'

opts = webdriver.ChromeOptions()
opts.add_argument('headless')
opts.add_argument('window-size=1920,1080')
driver = webdriver.Chrome(monitor_path+'chromedriver',chrome_options=opts)
try:
    product_urls = open(monitor_path+'products.txt', 'r').readlines()
    row =[today.strftime("%H:%M:%S")]
    changes = []
    column_idx = 1
    for product in product_urls:
        driver.get(product)

        price_xpath = "//div[contains(@class,'prdc_default_info')]//strong[@class='sale_price']"
        price_tag = driver.find_element_by_xpath(price_xpath)
        row.append(price_tag.text)

        title_tag = driver.find_element_by_xpath("//div[contains(@class,'heading')]/h2")
        column = string.ascii_uppercase[column_idx]
        worksheet[column+'1'] = title_tag.text

        prev_price = worksheet[column+'2'].value
        curr_price = price_tag.text
        if prev_price and prev_price != curr_price:
            changes.append((title_tag.text,prev_price,curr_price))

        worksheet[column+'2'] = price_tag.text
        column_idx += 1

    worksheet.append(row)
    result_xlsx.save(file_path)

    if changes:
        from my_email import send_mail
        contents = str(today)[:-7] + ': 가격이 변동된 상품이 있습니다.\n\n'
        
        for c in changes:
            contents += '(%s) %s => %s\n'%(c[0], c[1], c[2])

        send_mail('이태화', 'alghost.lee@gmail.com', contents)
   
except Exception as e:
    print(e)
finally:
    driver.quit()
