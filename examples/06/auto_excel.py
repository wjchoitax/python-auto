from os import listdir
from openpyxl import load_workbook, Workbook

files = listdir('.')

result_xlsx = Workbook()
result_sheet = result_xlsx.active

for myfile in files:
    if myfile[-4:] != 'xlsx':
        continue

    tg_xlsx = load_workbook(myfile, read_only=True)
    tg_sheet = tg_xlsx.active
    
    for row in tg_sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)

        result_sheet.append(row_data)

result_xlsx.save('result.xlsx')
