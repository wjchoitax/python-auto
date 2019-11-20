import openpyxl
import smtplib
import email
import pprint

def get_not_paid_members(ws):
    not_paid_list = []
    for row_index in range(2, ws.max_row + 1):
        if '미결제' in ws.cell(row_index, 4).value:
            not_paid_list.append(
                [ws.cell(row_index, 1).value,
                 ws.cell(row_index, 2).value])
    return not_paid_list


def send_email_for_ask_money(members):
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login('soongon', 'password')

    for member in members:
        # 텍스트 파일
        msg = email.mime.multipart.MIMEMultipart("alternative")
        msg["From"] = 'soongon@gmail.com'
        msg["To"] = member[1]
        msg["Subject"] = member[0] + '님 회비납부 요망'
        contents = '회비 부탁해요'
        text = email.mime.text.MIMEText(_text=contents, _charset="utf-8")
        msg.attach(text)
        smtp.sendmail('soongon@gmail.com', member[1], msg.as_string())
        print(member[0] + '님에게 메일 전송 완료')
    
    smtp.quit()

def report_ask_money_result_to_me():
    pass


def main():
    wb = openpyxl.load_workbook('수강생_결제정보.xlsx')
    ws = wb['결제정보']

    not_paid_members = get_not_paid_members(ws)

    send_email_for_ask_money(not_paid_members)

    report_ask_money_result_to_me()


main()
