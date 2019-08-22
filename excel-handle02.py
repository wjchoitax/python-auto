import openpyxl


def create_workbook_with_not_paid(not_paid_list):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('미결제 명단')
    #헤더추가
    for i, row in enumerate(not_paid_list):
        for j, cell in enumerate(row):
            sheet.cell(i+1, j+1, cell)

    wb.save('not-paid.xlsx')
    print('save ok..')


def main():
    wb = openpyxl.load_workbook('C:\\Users\\KITRI\\Desktop\\examples\\07\\수강생_결제정보.xlsx')
    sheet = wb['결제정보']

    not_paid_list = []
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row_index, 4).value == '미결제':
            not_paid_list.append([
                sheet.cell(row_index, 1).value,
                sheet.cell(row_index, 2).value,
                sheet.cell(row_index, 3).value,
                sheet.cell(row_index, 4).value,
                sheet.cell(row_index, 5).value
            ])

    create_workbook_with_not_paid(not_paid_list)


main()
